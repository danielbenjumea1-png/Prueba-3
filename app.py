import streamlit as st
import pandas as pd
import numpy as np
import easyocr
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from PIL import Image
import re

# ================================
# CONFIGURACIÓN DE COLORES
# ================================
COLOR_VERDE = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
COLOR_MORADO = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

st.title("📚 Inventario Biblioteca con IA OCR")
st.write("Escanea códigos con la cámara. Si existe se marca en verde, si no existe se agrega en morado.")

# ================================
# SUBIR ARCHIVO EXCEL
# ================================
uploaded_file = st.file_uploader("Sube tu archivo Excel del inventario", type=["xlsx"])
if uploaded_file:
    excel_path = "inventario.xlsx"
    with open(excel_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    wb = load_workbook(excel_path)
    sheet = wb.active
    df = pd.read_excel(excel_path)

    # Detectar columna que contiene la palabra "codigo"
    codigo_columna = None
    for col in df.columns:
        if "codigo" in col.lower():
            codigo_columna = col
            break

    if not codigo_columna:
        st.error("No se encontró ninguna columna que contenga 'codigo'.")
        st.stop()

    # Crear diccionario código → fila
    codigo_a_fila = {str(row[codigo_columna]).strip(): idx + 2 for idx, row in df.iterrows()}

    # ================================
    # OCR
    # ================================
    reader = easyocr.Reader(['es', 'en'])

    st.subheader("Escanea el código")
    img_file = st.camera_input("Toma una foto del código")
    codigo_manual = st.text_input("Ingresa el código manualmente si es necesario")

    codigo_detectado = None

    # ================================
    # PROCESAR IMAGEN CON OCR
    # ================================
    if img_file:
        img = Image.open(img_file)
        img_array = np.array(img)

        textos = reader.readtext(img_array, detail=0)

        # PALABRAS QUE SE DEBEN IGNORAR
        frases_prohibidas = [
            "sistemadeinformacionbibliografico",
            "sistemadeinformacion",
            "bibliografico",
            "biblioteca",
            "universidad",
            "cooperativa",
            "colombia"
        ]

        posibles_codigos = []

        for t in textos:
            t_limpio = t.lower().replace(" ", "").replace("-", "").strip()

            # ❌ Ignorar frases prohibidas institucionales
            if any(frase in t_limpio for frase in frases_prohibidas):
                continue

            # ✔ Detectar códigos que comienzan con B y tienen números después
            # Ejemplo: B0087034, B0451612, B0091274
            if re.fullmatch(r"b\d{6,8}", t_limpio):
                posibles_codigos.append(t_limpio.upper())
                continue

            # ✔ Alternativamente, detectar códigos alfanuméricos largos válidos
            if t_limpio.startswith("b") and len(t_limpio) >= 7:
                posibles_codigos.append(t_limpio.upper())

        if posibles_codigos:
            # Tomar el más largo (generalmente el correcto)
            codigo_detectado = max(posibles_codigos, key=len)
            st.success(f"Código detectado automáticamente: **{codigo_detectado}**")
        else:
            st.warning("No se encontró un código válido en la imagen. Usa la entrada manual.")

    # ================================
    # BOTÓN PARA ACTUALIZAR INVENTARIO
    # ================================
    if st.button("Actualizar Inventario"):

        # Prioridad 1: OCR
        if codigo_detectado:
            codigo = codigo_detectado.strip()
        # Prioridad 2: manual
        elif codigo_manual.strip() != "":
            codigo = codigo_manual.strip().upper()
        else:
            st.error("No se detectó ningún código. Ingresa uno manualmente.")
            st.stop()

        # ================================
        # ACTUALIZAR EXCEL
        # ================================
        if codigo in codigo_a_fila:
            fila = codigo_a_fila[codigo]
            celda = f"A{fila}"
            sheet[celda].fill = COLOR_VERDE
            sheet[celda].font = Font(bold=True)
            st.success(f"✔ Código {codigo} encontrado y marcado en verde.")
        else:
            nueva_fila = sheet.max_row + 1
            sheet[f"A{nueva_fila}"] = codigo
            sheet[f"A{nueva_fila}"].fill = COLOR_MORADO
            sheet[f"A{nueva_fila}"].font = Font(bold=True)
            st.warning(f"➕ Código {codigo} agregado como nuevo y marcado en morado.")

        wb.save(excel_path)

    # ================================
    # MOSTRAR INVENTARIO
    # ================================
    st.subheader("Inventario actualizado")
    st.dataframe(pd.read_excel(excel_path))

    with open(excel_path, "rb") as f:
        st.download_button("Descargar Excel actualizado", f, file_name="inventario_actualizado.xlsx")
